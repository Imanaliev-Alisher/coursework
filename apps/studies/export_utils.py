"""
Утилиты для экспорта расписания в различные форматы
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_pdf_timetable(timetable_data, title="Расписание"):
    """
    Генерирует PDF файл с расписанием
    
    Args:
        timetable_data: список словарей с данными расписания
        title: заголовок документа
    
    Returns:
        HttpResponse с PDF файлом
    """
    buffer = BytesIO()
    
    # Создаем документ в альбомной ориентации
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Стиль заголовка
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=1  # Center
    )
    
    # Добавляем заголовок
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    if not timetable_data:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            alignment=1
        )
        elements.append(Paragraph("Нет данных для отображения", no_data_style))
    else:
        # Подготовка данных для таблицы
        table_data = [
            ['День', 'Время', 'Предмет', 'Тип', 'Аудитория', 'Преподаватели', 'Группы', 'Неделя']
        ]
        
        for item in timetable_data:
            teachers_str = ', '.join(item.get('teachers', []))
            groups_str = ', '.join(item.get('groups', []))
            
            table_data.append([
                str(item.get('day', '')),
                str(item.get('time_slot', '')),
                str(item.get('subject', '')),
                str(item.get('subject_type', '')),
                str(item.get('audience', '')),
                teachers_str[:50] + '...' if len(teachers_str) > 50 else teachers_str,
                groups_str[:30] + '...' if len(groups_str) > 30 else groups_str,
                str(item.get('week_type', ''))
            ])
        
        # Создаем таблицу
        table = Table(table_data, colWidths=[3*cm, 3*cm, 4*cm, 2.5*cm, 3*cm, 4*cm, 3*cm, 2.5*cm])
        
        # Стиль таблицы
        table.setStyle(TableStyle([
            # Заголовок
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Данные
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 1), (-1, -1), 6),
            ('RIGHTPADDING', (0, 1), (-1, -1), 6),
            
            # Сетка
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Чередующиеся цвета строк
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))
        
        elements.append(table)
    
    # Строим PDF
    doc.build(elements)
    
    # Получаем значение буфера
    pdf = buffer.getvalue()
    buffer.close()
    
    # Создаем HTTP ответ
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable.pdf"'
    response.write(pdf)
    
    return response


def generate_excel_timetable(timetable_data, title="Расписание"):
    """
    Генерирует Excel файл с расписанием
    
    Args:
        timetable_data: список словарей с данными расписания
        title: заголовок документа
    
    Returns:
        HttpResponse с Excel файлом
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Расписание"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок документа
    worksheet.merge_cells('A1:H1')
    title_cell = worksheet['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 25
    
    # Заголовки столбцов
    headers = ['День недели', 'Время', 'Предмет', 'Тип', 'Аудитория', 'Преподаватели', 'Группы', 'Неделя']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    worksheet.row_dimensions[2].height = 30
    
    # Данные
    if not timetable_data:
        worksheet.merge_cells('A3:H3')
        no_data_cell = worksheet['A3']
        no_data_cell.value = "Нет данных для отображения"
        no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_data_cell.font = Font(italic=True, color="666666")
    else:
        for row_num, item in enumerate(timetable_data, 3):
            teachers_str = ', '.join(item.get('teachers', []))
            groups_str = ', '.join(item.get('groups', []))
            
            row_data = [
                item.get('day', ''),
                item.get('time_slot', ''),
                item.get('subject', ''),
                item.get('subject_type', ''),
                item.get('audience', ''),
                teachers_str,
                groups_str,
                item.get('week_type', '')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(value)
                cell.alignment = data_alignment
                cell.border = border
                
                # Чередующиеся цвета
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            worksheet.row_dimensions[row_num].height = 20
    
    # Установка ширины столбцов
    column_widths = [15, 20, 25, 15, 20, 30, 25, 15]
    for col_num, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Сохранение в буфер
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="timetable.xlsx"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response
